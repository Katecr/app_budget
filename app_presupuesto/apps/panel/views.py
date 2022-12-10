import math

from django.shortcuts import render
from django.http.response import HttpResponse
from django.views.generic import TemplateView


from openpyxl import Workbook, load_workbook

from .models import FormBudget


class IndexView(TemplateView):
	template_name = 'panel/index.html'

	def get(self, request, *args, **kwars):
		return render(request, self.template_name,{})


def list(request):
	if "GET" == request.method:
		return render(request, 'panel/list.html', {})
	else:
		excel_file = request.FILES["excel_file"]

		# you may put validations here to check extension or file size
		wbk = load_workbook(excel_file)

		# getting a particular sheet by name out of many sheets
		worksheet = wbk["Hoja1"]

		excel_data = []
		# iterating over the rows and
		# getting value from each cell in row
		for row in worksheet.iter_rows():
			for row in worksheet.iter_rows():
				row_data = []
				for cell in row:
					row_data.append(str(cell.value))
				excel_data.append(row_data)

		sueldo_anio = worksheet['A2'].value
		salario = worksheet['B2'].value
		cesantias = worksheet['C2'].value
		prima = worksheet['D2'].value
		int_cesantias = worksheet['E2'].value
		vacaciones = worksheet['F2'].value
		total_prest = worksheet['G2'].value
		salud = worksheet['H2'].value
		pension = worksheet['I2'].value
		arl = worksheet['J2'].value
		comfama = worksheet['K2'].value
		sena = worksheet['L2'].value
		icbf = worksheet['M2'].value
		total_aportes = worksheet['N2'].value
		tipo_presupuesto = worksheet['O2'].value

		wb = Workbook()
		ws = wb.active

		ws["A1"] = "Enero"
		ws["B1"] = "Febrero"
		ws["C1"] = "Marzo"
		ws["D1"] = "Abril"
		ws["E1"] = "Mayo"
		ws["F1"] = "Junio"
		ws["G1"] = "Julio"
		ws["H1"] = "Agosto"
		ws["I1"] = "Septiembre"
		ws["J1"] = "Octubre"
		ws["K1"] = "Noviembre"
		ws["L1"] = "Diciembre"

		cont = 2

	

		if tipo_presupuesto == 51 or tipo_presupuesto == '51':
			for i in range(1,15):
				if i == 1:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = salario / 12
					cont += 1
				elif i == 2:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = ''
					cont += 1
				elif i == 3:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = ''
					cont += 1
				elif i == 4:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = ''
					cont += 1
				elif i == 5:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = cesantias / 12
					cont += 1
				elif i == 6:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = int_cesantias / 12
					cont += 1
				elif i == 7:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = prima / 12
					cont += 1
				elif i == 8:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = vacaciones / 12
					cont += 1
				elif i == 9:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = ''
					cont += 1
				elif i == 10:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = arl / 12
					cont += 1
				elif i == 11:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = salud / 12
					cont += 1
				elif i == 12:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = pension / 12
					cont += 1
				elif i == 13:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = comfama / 12
					cont += 1
				elif i == 14:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = icbf / 12
					cont += 1
				elif i == 15:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = sena/12
					cont += 1
		elif tipo_presupuesto == 72 or tipo_presupuesto == '72':
			for i in range(1,14):
				if i == 1:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = salario / 12
					cont += 1
				elif i == 2:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = ''
					cont += 1
				elif i == 3:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = cesantias / 12
					cont += 1
				elif i == 4:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = int_cesantias / 12
					cont += 1
				elif i == 5:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = prima / 12
					cont += 1
				elif i == 6:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = vacaciones / 12
					cont += 1
				elif i == 7:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = ''
					cont += 1
				elif i == 8:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = ''
					cont += 1
				elif i == 9:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = arl / 12
					cont += 1
				elif i == 10:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = salud / 12
					cont += 1
				elif i == 11:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = pension / 12
					cont += 1
				elif i == 12:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = comfama / 12
					cont += 1
				elif i == 13:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = icbf / 12
					cont += 1
				elif i == 14:
					for j in range(1,13):
						ws.cell(row = cont, column = j).value = sena / 12
					cont += 1

		name_file = "ReportePresupuesto.xlsx"
		response = HttpResponse(content_type = 'application/ms-excel')
		content = "attachment; filename = {0}".format(name_file)
		response['Content-Disposition'] = content
		wb.save(response)
		return response

		# return render(request, 'panel/list.html', {"excel_data":excel_data})


class ReporteExcelView(TemplateView):

	def get(self, request, *args, **kwars):
		# budgets = FormBudget.objects.all()
		wb = Workbook()
		ws = wb.active
		
		# para concatenas
		# ws.merge_cells('B1:E1')

		# Nombres de emcabezados quemados
		ws["A1"] = "Enero"
		ws["B1"] = "Febrero"
		ws["C1"] = "Marzo"
		ws["D1"] = "Abril"
		ws["E1"] = "Mayo"
		ws["F1"] = "Junio"
		ws["G1"] = "Julio"
		ws["H1"] = "Agosto"
		ws["I1"] = "Septiembre"
		ws["J1"] = "Octubre"
		ws["K1"] = "Noviembre"
		ws["L1"] = "Diciembre"
	

		cont = 2

		for budget in range(1,4):
			ws.cell(row = cont, column = 1).value = budget
			ws.cell(row = cont, column = 2).value = 'prueba'
			ws.cell(row = cont, column = 3).value = 'prueba'
			ws.cell(row = cont, column = 4).value = 'prueba'
			cont += 1

		name_file = "ReportePresupuesto.xlsx"
		response = HttpResponse(content_type = 'application/ms-excel')
		content = "attachment; filename = {0}".format(name_file)
		response['Content-Disposition'] = content
		wb.save(response)
		return response
